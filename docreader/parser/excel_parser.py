"""
Excel Parser Module

This module provides functionality to parse Excel files (.xlsx, .xls) into
structured Document objects with text content and chunks. It supports multiple
sheets and handles various Excel formats using pandas.
"""
import logging
from io import BytesIO
from typing import List

import pandas as pd

from docreader.models.document import Chunk, Document
from docreader.parser.base_parser import BaseParser
from docreader.parser.excel_convert import (
    convert_excel_to_xlsx_bytes,
    detect_excel_format,
    engine_for_format,
    normalize_excel_bytes,
)
from docreader.parser.xlsx_merge import fill_merged_cells_xlsx
from docreader.parser.xlsx_repair import repair_xlsx_bytes

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel files (.xlsx, .xls).
    
    This parser extracts text content from Excel files by processing all sheets
    and converting each row into a structured text format. Each row becomes a
    separate chunk with key-value pairs.
    
    Features:
        - Supports multiple sheets in a single Excel file
        - Automatically removes completely empty rows
        - Converts each row to "column: value" format
        - Creates individual chunks for each row for better granularity
        
    Example:
        >>> parser = ExcelParser()
        >>> with open("data.xlsx", "rb") as f:
        ...     content = f.read()
        ...     document = parser.parse_into_text(content)
        >>> print(document.content)
        Name: John,Age: 30,City: NYC
        Name: Jane,Age: 25,City: LA
    """
    
    def parse_into_text(self, content: bytes) -> Document:
        """Parse Excel file bytes into a Document object.
        
        Args:
            content: Raw bytes of the Excel file
            
        Returns:
            Document: Parsed document containing:
                - content: Full text with all rows from all sheets
                - chunks: List of Chunk objects, one per row
                
        Note:
            - Empty rows (all NaN values) are automatically skipped
            - Each row is formatted as: "col1: val1,col2: val2,..."
            - Chunks maintain sequential ordering across all sheets
        """
        chunks: List[Chunk] = []
        text: List[str] = []
        start, end = 0, 0

        excel_file = _open_excel_file(content, file_type=self.file_type)
        
        # Process each sheet in the Excel file
        for excel_sheet_name in excel_file.sheet_names:
            df = _read_sheet_dataframe(excel_file, excel_sheet_name)
            # Remove rows where all values are NaN (completely empty rows)
            df.dropna(how="all", inplace=True)

            # Process each row in the DataFrame
            for _, row in df.iterrows():
                page_content = []
                # Build key-value pairs for non-null values
                for k, v in row.items():
                    if pd.notna(v):  # Skip NaN/null values
                        page_content.append(f"{k}: {v}")
                
                # Skip rows with no valid content
                if not page_content:
                    continue
                
                # Format row as comma-separated key-value pairs
                content_row = ",".join(page_content) + "\n"
                end += len(content_row)
                text.append(content_row)
                
                # Create a chunk for this row with position tracking
                chunks.append(
                    Chunk(content=content_row, seq=len(chunks), start=start, end=end)
                )
                start = end

        # Combine all text and return as Document
        return Document(content="".join(text), chunks=chunks)


def _read_sheet_dataframe(excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    """Read a worksheet into a DataFrame with stable column labels."""
    from openpyxl.utils import get_column_letter

    # XLSX is preprocessed (merge fill); use A/B/C column letters and keep row 1 as data.
    if excel_file.engine == "openpyxl":
        df = excel_file.parse(sheet_name=sheet_name, header=None)
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
        return df

    df = excel_file.parse(sheet_name=sheet_name, header=0)
    if df.empty:
        df = excel_file.parse(sheet_name=sheet_name, header=None)
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
    elif any(str(col).startswith("Unnamed:") for col in df.columns):
        df = excel_file.parse(sheet_name=sheet_name, header=None)
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
    return df


def _prepare_xlsx_bytes(data: bytes) -> bytes:
    repaired = repair_xlsx_bytes(data)
    if repaired is not None:
        data = repaired
    return fill_merged_cells_xlsx(data)


def _open_excel_file(content: bytes, file_type: str | None = None) -> pd.ExcelFile:
    """Open an Excel workbook with explicit engine selection and fallbacks."""
    data = content
    converted_via_soffice = False

    while True:
        ext = detect_excel_format(data)
        if ext is None:
            if converted_via_soffice:
                raise ValueError(
                    "Excel file format cannot be determined, you must specify an engine manually."
                )
            try:
                data = normalize_excel_bytes(data, file_type=file_type)
            except ValueError as exc:
                raise ValueError(
                    "Excel file format cannot be determined, you must specify an engine manually."
                ) from exc
            converted_via_soffice = True
            continue

        if ext == "ods":
            converted = convert_excel_to_xlsx_bytes(data, suffix=".ods")
            if converted:
                data = converted
                continue

        engine = engine_for_format(ext)
        if ext == "xlsx":
            data = _prepare_xlsx_bytes(data)
            engine = "openpyxl"
        try:
            return pd.ExcelFile(BytesIO(data), engine=engine)
        except ImportError as exc:
            raise ValueError(
                f"Excel engine {engine!r} is not available for .{ext} files"
            ) from exc
        except KeyError as exc:
            if "sharedStrings.xml" not in str(exc) or engine != "openpyxl":
                raise
            repaired = repair_xlsx_bytes(data)
            if repaired is None:
                raise
            logger.info("Repaired XLSX sharedStrings packaging before parse")
            data = _prepare_xlsx_bytes(repaired)
            continue
        except ValueError as exc:
            if converted_via_soffice or "cannot be determined" not in str(exc):
                raise
            try:
                data = normalize_excel_bytes(content, file_type=file_type)
            except ValueError:
                raise
            converted_via_soffice = True
            continue


if __name__ == "__main__":
    # Example usage: Parse an Excel file and display results
    logging.basicConfig(level=logging.DEBUG)

    # Specify the path to your Excel file
    your_file = "/path/to/your/file.xlsx"
    parser = ExcelParser()
    
    # Read and parse the Excel file
    with open(your_file, "rb") as f:
        content = f.read()
        document = parser.parse_into_text(content)
        
        # Display the full document content
        logger.error(document.content)

        # Display the first chunk as an example
        for chunk in document.chunks:
            logger.error(chunk.content)
            break  # Only show the first chunk
