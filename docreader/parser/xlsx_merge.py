"""Fill merged cell values before pandas reads an XLSX workbook."""

from __future__ import annotations

import logging
import zipfile
from io import BytesIO

logger = logging.getLogger(__name__)


def fill_merged_cells_xlsx(content: bytes) -> bytes:
    """Unmerge ranges and copy the master cell value into every covered cell.

    openpyxl only stores values on the top-left cell of a merge; pandas then
    sees NaN in the rest. Filling makes row-wise RAG chunks retain context.
    """
    if not zipfile.is_zipfile(BytesIO(content)):
        return content

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), data_only=True)
    changed = False
    for ws in wb.worksheets:
        if not ws.merged_cells.ranges:
            continue
        for merge_range in list(ws.merged_cells.ranges):
            master_value = ws.cell(merge_range.min_row, merge_range.min_col).value
            ws.unmerge_cells(str(merge_range))
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    ws.cell(row, col).value = master_value
            changed = True

    if not changed:
        return content

    out = BytesIO()
    wb.save(out)
    logger.info("Filled merged cells in XLSX before parse")
    return out.getvalue()
